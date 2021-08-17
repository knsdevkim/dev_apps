from django.shortcuts import render, HttpResponse
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.decorators import login_required
from django.views.generic import CreateView, ListView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db import IntegrityError
from openpyxl import load_workbook

from datetime import datetime
import re

from .forms import (
    UserCreateForm,
    LoginUserForm,
    UserUpdateForm,
    SyncForm)
from .models import *


def split_racklocation(location):
    location = re.split('[A-Z]+', location)
    location.remove('')
    return location

# Create your views here.


@login_required
def panel(request):
    context = {}
    context['context_page_title']  = 'Dashboard'
    context['context_page_active'] = '<i class="fa fa-home"></i>'
    return render(request, 'components/staticview/dashboard.html', context = context)


class LoginViewUser(LoginView):
    template_name       = 'components/credentials/login.html'
    authentication_form = LoginUserForm


class ListViewUsers(LoginRequiredMixin, ListView):
    model               = Users
    template_name       = 'components/crud/list.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_page_title']  = 'WMS Users'
        kwargs['context_page_active'] = '<i class="fa fa-users"></i>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return self.model.objects.all()


class CreateViewUser(LoginRequiredMixin, CreateView):
    template_name = 'components/crud/input.html'
    form_class    = UserCreateForm

    def get_context_data(self, **kwargs):
        kwargs['context_page_title']  = 'New User'
        kwargs['context_page_active'] = '<i class="fa fa-user"></i>'
        kwargs['context_button']      = 'Create Admin'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'New admin is created.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Could not create see errors below and try again.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('app:createadmin')


class UpdateViewUser(LoginRequiredMixin, UpdateView):
    template_name = 'components/crud/input.html'
    form_class    = UserUpdateForm

    def get_context_data(self, **kwargs):
        kwargs['context_page_title']  = 'Update User'
        kwargs['context_page_active'] = '<i class="fa fa-user"></i>'
        kwargs['context_button']      = 'Update Information'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Users.objects.filter(pk = self.kwargs['pk'])

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'Successfully updated.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Could not update see errors below and try again.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('app:updateuser', kwargs = {
            'pk': self.kwargs['pk']
        })

class DeleteViewUser(LoginRequiredMixin, DeleteView):
    model         = Users
    template_name = None

    def get_success_url(self):
        messages.success(self.request, 'User has been remove.')
        return reverse_lazy('app:users')


class LogoutViewUser(LogoutView):
    pass

class FormViewSync(LoginRequiredMixin, FormView):
    template_name = 'components/crud/input.html'
    form_class    = SyncForm

    def get_context_data(self, **kwargs):
        kwargs['context_page_title']  = 'Sync WMS'
        kwargs['context_page_active'] = '<i class="fa fa-upload"></i>'
        kwargs['context_button']      = 'Sync'
        return super().get_context_data(**kwargs)

    def post(self, request, *args, **kwargs):
        return self.form_valid(self.get_form()) if self.get_form().is_valid() else self.form_invalid(self.get_form())

    def form_valid(self, form):
        try:
            xlsx_workbook   = load_workbook(filename = self.request.FILES.get('fileinput', ''), read_only = True)
            xlsx_sheet      = xlsx_workbook['Sheet1']
            model_folder = Folder.objects.create(
                pk         = None,
                foldername = self.request.POST.get('foldername', '')
            )
            count = 0
            for cell_buffer in range(2, xlsx_sheet.max_row):
                count += 1
                # MATCH ONLY THIS LOCATION EXPRESSION TO GET THE DATA (SORT)
                if re.findall(r'\AR(\d{1,2})L(\d{1,2})C(\d{1,2})\Z', str(xlsx_sheet['G' + str(cell_buffer)].value)):
                    xlsx_racklocation  = split_racklocation(xlsx_sheet['G' + str(cell_buffer)].value)
                    model_racklocation = Racklocation.objects.create(
                        pk       = None,
                        rack     = xlsx_racklocation[0],
                        location = xlsx_racklocation[1],
                        column   = xlsx_racklocation[2]
                    )
                    model_product = Products(
                        pk              = None,
                        folder          = model_folder,
                        stockcode       = xlsx_sheet['A' + str(cell_buffer)].value,
                        description     = xlsx_sheet['B' + str(cell_buffer)].value,
                        case            = xlsx_sheet['C' + str(cell_buffer)].value,
                        inner_box       = xlsx_sheet['D' + str(cell_buffer)].value,
                        piece           = xlsx_sheet['E' + str(cell_buffer)].value,
                        inventory_value = xlsx_sheet['F' + str(cell_buffer)].value,
                        batch_code      = xlsx_sheet['H' + str(cell_buffer)].value,
                        expiry_date     = datetime.strptime(xlsx_sheet['I' + str(cell_buffer)].value, "%m/%d/%Y")
                    ).save()
                    model_product.racklocation.set(model_racklocation)
            messages.success(self.request, 'Successfully synced.')
        except IntegrityError as e:
            code, message = e.args
            if code == 1062:
                messages.error(self.request, 'Folder name is already exists.')
            else:
                messages.error(self.request, f'<strong>Backend Error:</strong> {e.args}.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Could not sync see errors below and try again.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('app:sync')
