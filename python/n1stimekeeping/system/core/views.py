import time
from datetime import datetime, time as date_time

from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import (
    FormView
)

from .forms import (
    SyncFileForm,
    TimeLogForm
)
from .models import (
    EmployeeRecordModel,
    EmployeeInformationModel,
    TimeKeepingDateModel,
    TimeLogsModel
)


# Create your views here.
class SyncFileClassBaseView(FormView):
    template_name = 'components/crud/create.html'
    form_class = SyncFileForm

    def form_valid(self, form):
        from openpyxl import load_workbook

        xls_workbook = load_workbook(filename=self.request.FILES.get('file', ''), read_only=True)
        xls_sheet = xls_workbook['Sheet1']

        for i in range(2, xls_sheet.max_row):
            # EMPLOYEE RECORD MODEL
            emp_fk = EmployeeRecordModel.objects.create(
                pk=None,
                employee=xls_sheet['A' + str(i)].value,
                business_unit=xls_sheet['B' + str(i)].value,
                first_name=xls_sheet['C' + str(i)].value,
                last_name=xls_sheet['D' + str(i)].value,
                middle_name=xls_sheet['E' + str(i)].value,
                mobile_number=xls_sheet['F' + str(i)].value,
                email_address=xls_sheet['G' + str(i)].value,
                employee_position=xls_sheet['L' + str(i)].value,
                employee_status=xls_sheet['O' + str(i)].value,
                date_hired=xls_sheet['K' + str(i)].value,
                date_regularized=xls_sheet['P' + str(i)].value
            )

            # EMPLOYEE INFORMATION MODEL
            EmployeeInformationModel.objects.create(
                pk=None,
                employee=emp_fk,
                civil_status=xls_sheet['I' + str(i)].value,
                birthdate=xls_sheet['J' + str(i)].value,
                address_1=xls_sheet['M' + str(i)].value,
                address_2=xls_sheet['N' + str(i)].value,
                third_party_employer=xls_sheet['Q' + str(i)].value,
                shirt_size=xls_sheet['R' + str(i)].value,
                active_status=xls_sheet['W' + str(i)].value
            )

        return super(SyncFileClassBaseView, self).form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Something went wrong in the form to process, Try again.')
        return super(SyncFileClassBaseView, self).form_invalid(form)

    def post(self, request, *args, **kwargs):
        return self.form_valid(self.get_form()) if self.get_form().is_valid() else self.form_invalid(self.get_form())

    def get_success_url(self):
        return reverse_lazy('core:syncfile')


class TimeKeepingClassBaseView(FormView):
    template_name = 'components/view/timekeeping.html'
    form_class = TimeLogForm

    def get_context_data(self, **kwargs):
        kwargs['context_data'] = TimeLogsModel.objects.all()
        return super(TimeKeepingClassBaseView, self).get_context_data(**kwargs)

    def form_valid(self, form):
        employee_id_scanner = self.request.POST.get('employee_id', '').replace('AE', '')

        try:
            employee_fk = EmployeeRecordModel.objects.get(employee=f'AE{employee_id_scanner}')

            def set_time(fk):
                user_time = datetime.strptime(self.request.POST['time'], '%I:%M:%S %p').time()
                start = date_time(5, 0)
                end = date_time(17, 0)
                breaks = [
                    (date_time(10, 0), date_time(10, 10)),
                    (date_time(12, 0), date_time(12, 55)),
                    (date_time(15, 0), date_time(15, 10))
                ]

                if not (start <= user_time <= end):
                    TimeLogsModel.objects.create(
                        date=fk,
                        time=datetime.strptime(self.request.POST['time'], '%I:%M:%S %p').strftime('%I:%M %p'),
                        status='TIME OUT'
                    )
                    return 'TIME OUT'
                else:
                    for start_break, end_break in breaks:
                        if start_break <= user_time <= end_break:
                            TimeLogsModel.objects.create(
                                date=fk,
                                time=datetime.strptime(self.request.POST['time'], '%I:%M:%S %p').strftime('%I:%M %p'),
                                status='BREAK OUT'
                            )
                            return 'BREAK OUT'

                TimeLogsModel.objects.create(
                    date=fk,
                    time=datetime.strptime(self.request.POST['time'], '%I:%M:%S %p').strftime('%I:%M %p'),
                    status='TIME IN'
                )

                return 'TIME IN'

            if TimeKeepingDateModel.objects.filter(employee=employee_fk, date=time.strftime("%Y-%m-%d")).exists():
                date_fk = TimeKeepingDateModel.objects.get(employee=employee_fk, date=time.strftime('%Y-%m-%d'))

                get_latest = TimeLogsModel.objects.filter(date=date_fk).latest()

                status = set_time(date_fk)

                messages.success(self.request,
                                 f'Hello, {employee_fk.employee} - {employee_fk.first_name}, {employee_fk.last_name}. You are now {status}.')
            else:
                date_fk = TimeKeepingDateModel.objects.create(employee=employee_fk)
                status = set_time(date_fk)
                messages.success(self.request,
                                 f'Hello, {employee_fk.employee} - {employee_fk.first_name}, {employee_fk.last_name}. You are now {status}.')

        except EmployeeRecordModel.DoesNotExist:
            messages.error(self.request, f'No records found in AE{employee_id_scanner}.')

        return super(TimeKeepingClassBaseView, self).form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Something went wrong, Try again.')
        return super(TimeKeepingClassBaseView, self).form_invalid(form)

    def post(self, request, *args, **kwargs):
        return self.form_valid(self.get_form()) if self.get_form().is_valid() else self.form_invalid(self.get_form())

    def get_success_url(self):
        return reverse_lazy('core:timekeeping')
