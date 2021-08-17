from django.shortcuts import render, HttpResponse
from django.urls import reverse_lazy
from django.db import IntegrityError
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib.auth.views import (
    LoginView, LogoutView
)
from django.views.generic import *


from openpyxl import load_workbook


from .forms import *
from .models import *
from .decorators import *


# Create your views here.


@login_required
@unauthenticated_user
def dashboard(request):
    return render(request, 'components/dashboard.html', context = {})

@login_required
@allowed_users(allowed_roles = ['admin'])
def approve_account(request, pk):
    user = Users.objects.get(pk = pk)
    user.account_status = 'active'
    user.save()
    messages.success(request, 'Account approved!')
    return redirect('apps:accountlist')


@method_decorator(authenticated_user, name = 'dispatch')
class AuthenticateLoginView(LoginView):
    template_name       = 'components/authentication/login.html'
    authentication_form = LoginForm


class AuthenticateLogoutView(LogoutView):
    pass


@method_decorator(allowed_users(allowed_roles = ['admin']), name = 'dispatch')
class AccountListView(LoginRequiredMixin, ListView):
    template_name       = 'components/crud/read.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Account List'
        kwargs['context_icon']  = '<em class="fa fa-users"></em>'
        kwargs['context_links'] = f'<a href="{ reverse_lazy("apps:newaccount") }" class="btn btn-md btn-primary"><i class="fa fa-plus"></i> Add New Account</a><hr>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Users.objects.all()


@method_decorator(allowed_users(allowed_roles = ['admin']), name = 'dispatch')
class AccountCreateView(LoginRequiredMixin, CreateView):
    template_name = 'components/crud/form.html'
    form_class    = AccountForm

    def get_context_data(self, **kwargs):
        kwargs['context_title']        = 'New Account'
        kwargs['context_icon']         = '<em class="fa fa-plus"></em>'
        kwargs['context_button_title'] = 'Create Account'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        try:
            self.object = form.save()
            messages.success(self.request, '<em class="fa fa-check"></em> Account created!')
        except IntegrityError as e:
            messages.error(self.request, f'<em class="fa fa-exclamation"></em> {e}')
            return self.form_invalid(form)
        return super().form_valid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('apps:newaccount')


@method_decorator(allowed_users(allowed_roles = ['admin']), name = 'dispatch')
class AccountDeleteView(LoginRequiredMixin, DeleteView):
    model         = Users
    template_name = None

    def get_success_url(self):
        messages.success(self.request, '<em class="fa fa-check"></em> Account has been removed.')
        return reverse_lazy('apps:accountlist')


@method_decorator(allowed_users(allowed_roles = ['admin']), name = 'dispatch')
class SyncFileFormView(LoginRequiredMixin, FormView):
    template_name = 'components/crud/form.html'
    form_class    = SyncFileForm

    def get_context_data(self, **kwargs):
        kwargs['context_enctype']      = 'enctype="multipart/form-data"'
        kwargs['context_title']        = 'Sync File'
        kwargs['context_icon']         = '<em class="fa fa-upload"></em>'
        kwargs['context_button_title'] = 'Sync'
        return super().get_context_data(**kwargs)

    def post(self, request, *args, **kwargs):
        return self.form_valid(self.get_form()) if self.get_form().is_valid() else self.form_invalid(self.get_form())

    def form_valid(self, form):
        try:
            xlsx       = load_workbook(filename = self.request.FILES.get('filename', ''), read_only = True)
            xlsx_sheet = xlsx['Sheet1']
            mcp_model  = Mcpmasterlist()
            if self.request.POST.get('refresh_data', False):
                mcp_model.truncate()
            visual_count = 0
            for cell_number in range(2, xlsx_sheet.max_row):
                if xlsx_sheet['A' + str(cell_number)].value != '':
                    mcp_model.pk             = None
                    mcp_model.cust_code      = xlsx_sheet['A' + str(cell_number)].value
                    mcp_model.customer       = xlsx_sheet['B' + str(cell_number)].value
                    mcp_model.scode          = xlsx_sheet['C' + str(cell_number)].value
                    mcp_model.salesperson    = xlsx_sheet['D' + str(cell_number)].value
                    mcp_model.ave_nps        = xlsx_sheet['E' + str(cell_number)].value
                    mcp_model.class_label    = xlsx_sheet['F' + str(cell_number)].value
                    mcp_model.address        = xlsx_sheet['G' + str(cell_number)].value
                    mcp_model.area           = xlsx_sheet['H' + str(cell_number)].value
                    mcp_model.odd_even       = xlsx_sheet['I' + str(cell_number)].value
                    mcp_model.branch         = xlsx_sheet['J' + str(cell_number)].value
                    mcp_model.channel        = xlsx_sheet['K' + str(cell_number)].value
                    mcp_model.freq           = xlsx_sheet['L' + str(cell_number)].value
                    mcp_model.day            = xlsx_sheet['M' + str(cell_number)].value
                    mcp_model.cterm          = xlsx_sheet['N' + str(cell_number)].value
                    mcp_model.climit         = xlsx_sheet['O' + str(cell_number)].value
                    mcp_model.sman_type      = xlsx_sheet['P' + str(cell_number)].value
                    mcp_model.gtm            = xlsx_sheet['Q' + str(cell_number)].value
                    mcp_model.group          = xlsx_sheet['R' + str(cell_number)].value
                    mcp_model.town           = xlsx_sheet['S' + str(cell_number)].value
                    mcp_model.zip_code       = xlsx_sheet['T' + str(cell_number)].value
                    mcp_model.channel_group  = xlsx_sheet['U' + str(cell_number)].value
                    mcp_model.channel_group2 = xlsx_sheet['V' + str(cell_number)].value
                    mcp_model.chain          = xlsx_sheet['W' + str(cell_number)].value
                    mcp_model.area_class     = xlsx_sheet['X' + str(cell_number)].value
                    mcp_model.old_new        = xlsx_sheet['Y' + str(cell_number)].value
                    mcp_model.geolocation    = xlsx_sheet['Z' + str(cell_number)].value
                    mcp_model.save()
                    visual_count += 1
                    print(f'\rREAD LINE SAVED: {visual_count} :: CUST CODE: {xlsx_sheet["A" + str(cell_number)].value}', end = ' ', flush = True)
            messages.success(self.request, '<em class="fa fa-check"></em> New mcp masterlist data synced!')
        except Exception as e:
            messages.error(self.request, f'<em class="fa fa-exclamation"></em> <b>Exception error:</b> {e}')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, '<em class="fa fa-exclamation"></em> There is an error while syncing mcp master file.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('apps:syncfile')


@method_decorator(allowed_users(allowed_roles = ['admin']), name = 'dispatch')
class GtmListView(LoginRequiredMixin, ListView):
    template_name       = 'components/crud/read.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'GTM List'
        kwargs['context_icon']  = '<em class="fa fa-users"></em>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Mcpmasterlist.objects.all().values('gtm').distinct()


@method_decorator(allowed_users(allowed_roles = ['admin']), name = 'dispatch')
class SalesListView(LoginRequiredMixin, ListView):
    template_name       = 'components/crud/read.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Salesperson List'
        kwargs['context_icon']  = '<em class="fa fa-users"></em>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Mcpmasterlist.objects.filter(gtm = self.kwargs['gtm']).values('salesperson').distinct()


class CustomerListView(LoginRequiredMixin, ListView):
    template_name       = 'components/crud/read.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Customer List'
        kwargs['context_icon']  = '<em class="fa fa-users"></em>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Mcpmasterlist.objects.filter(salesperson = self.kwargs['salesperson'])


class CustomerInfoDetailView(LoginRequiredMixin, DetailView):
    template_name       = 'components/crud/detail.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Customer Details'
        kwargs['context_icon']  = '<em class="fa fa-user"></em>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Mcpmasterlist.objects.filter(pk = self.kwargs[self.pk_url_kwarg])


class AssignedGtmListView(LoginRequiredMixin, ListView):
    template_name       = 'components/crud/read.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Salesperson List'
        kwargs['context_icon']  = '<em class="fa fa-users"></em>'
        kwargs['context_links'] = '<a href="" class="btn btn-md btn-primary"><em class="fa fa-plus"></em> Assign</a><hr/>'
        return super().get_context_data(**kwargs)

    def get_queryset(self):
        return Mcpmasterlist.objects.filter(account_gtm = self.request.user.pk)
