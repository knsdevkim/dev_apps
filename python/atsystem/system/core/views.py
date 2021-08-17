from django.shortcuts import render, HttpResponse, HttpResponseRedirect, Http404
from django.urls import reverse_lazy
from django.contrib import messages
from django.views.generic import (
    FormView,
    ListView,
    TemplateView,
    UpdateView,
    DeleteView
)

from .forms import *
from .models import *

# Create your views here.

def generate_excel(request, folder):
    from django.conf import settings
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import os

    objects = ItEquipmentModel.objects.filter(folder = folder, is_claim = True)

    file_target = os.path.join(settings.MEDIA_ROOT, 'outputs/n1s_assettagging_output.xls')

    if os.path.exists(file_target):
        os.remove(file_target)

    excel_workbook = Workbook()
    excel_sheet = excel_workbook.active

    excel_sheet.title = 'Sheet1'

    excel_style_font = Font(name = 'Arial', bold = True, color = '00FFFFFF')
    excel_style_fill = PatternFill(fill_type = 'solid', fgColor = '00808000')
    excel_style_alignment = Alignment(horizontal = 'center', vertical = 'center')

    excel_sheet['A1'].value = 'FROM'
    excel_sheet['A1'].font = excel_style_font
    excel_sheet['A1'].fill = excel_style_fill
    excel_sheet['A1'].alignment = excel_style_alignment

    excel_sheet['B1'].value = 'TO'
    excel_sheet['B1'].font = excel_style_font
    excel_sheet['B1'].fill = excel_style_fill
    excel_sheet['B1'].alignment = excel_style_alignment

    excel_sheet['C1'].value = 'DEVICE'
    excel_sheet['C1'].font = excel_style_font
    excel_sheet['C1'].fill = excel_style_fill
    excel_sheet['C1'].alignment = excel_style_alignment

    excel_sheet['D1'].value = 'MOBILE NO.'
    excel_sheet['D1'].font = excel_style_font
    excel_sheet['D1'].fill = excel_style_fill
    excel_sheet['D1'].alignment = excel_style_alignment

    excel_sheet['E1'].value = 'IMEI' 
    excel_sheet['E1'].font = excel_style_font
    excel_sheet['E1'].fill = excel_style_fill
    excel_sheet['E1'].alignment = excel_style_alignment

    excel_sheet['F1'].value = 'ICCID'
    excel_sheet['F1'].font = excel_style_font
    excel_sheet['F1'].fill = excel_style_fill
    excel_sheet['F1'].alignment = excel_style_alignment

    excel_sheet['G1'].value = 'ACCESSORIES'  
    excel_sheet['G1'].font = excel_style_font
    excel_sheet['G1'].fill = excel_style_fill
    excel_sheet['G1'].alignment = excel_style_alignment

    excel_row_count = 1

    for field in objects:
        excel_row_count += 1

        excel_sheet['A' + str(excel_row_count)].value = field.userfrom
        excel_sheet['B' + str(excel_row_count)].value = field.custodian
        excel_sheet['C' + str(excel_row_count)].value = field.device
        excel_sheet['D' + str(excel_row_count)].value = field.mobile
        excel_sheet['E' + str(excel_row_count)].value = field.imei
        excel_sheet['F' + str(excel_row_count)].value = field.iccid
        excel_sheet['G' + str(excel_row_count)].value = field.accessories

    excel_workbook.save(file_target) 

    with open(file_target, 'rb') as file_read:
        response = HttpResponse(file_read.read(), 'application/vnd.ms-excel')
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_target)
        return response

    raise Http404

def claim(request):
    next = request.POST.get('next', '/')

    objects = ItEquipmentModel.objects.get(pk = request.POST.get('pk', 0))
    
    claim = True

    if objects.is_claim:
        claim = False

    objects.is_claim = claim
    objects.save()

    messages.success(request, f'{objects.custodian} claim modified.')

    return HttpResponseRedirect(next)

def itprint(request, pk):
    context = {}
    context['object'] = ItEquipmentModel.objects.filter(folder = pk)
    return render(request, 'print/it.html', context)

def adaprint(request, pk):
    context = {}
    context['object'] = AdaModel.objects.filter(folder = pk)
    return render(request, 'print/ada.html', context)

def getitprint(request, pk):
    context = {}
    context['object'] = ItEquipmentModel.objects.filter(pk = pk)
    return render(request, 'print/it.html', context)

def getadaprint(request, pk):
    context = {}
    context['object'] = AdaModel.objects.filter(pk = pk)
    return render(request, 'print/ada.html', context)

class DashboardClassBaseView(TemplateView):
    template_name = 'components/dashboard/view.html'

class UploadClassBaseView(FormView):
    template_name = 'components/crud/upload.html'
    form_class = UploadForm

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Upload File'
        kwargs['form'] = self.get_form()
        return super(UploadClassBaseView, self).get_context_data(**kwargs)

    def post(self, request, *args, **kwrags):
        form = self.get_form()
        return self.form_valid(form) if form.is_valid() else self.form_invalid(form)

    def form_valid(self, form):
        foldername = self.request.POST['folder']
        type_record = self.request.POST['type_record']

        from openpyxl import load_workbook

        xlsfile = load_workbook(filename = self.request.FILES['file'], read_only = True)
        sheet = xlsfile['Sheet1']

        folderfk = FolderModel.objects.create(
            folder = foldername,
            type_record = type_record
        )

        if type_record == 'ada':
            sync = AdaModel()
            for lineread in range(2, (sheet.max_row + 1)):
                if sheet['A' + str(lineread)].value != None:
                    sync.pk = None
                    sync.folder = folderfk
                    sync.area = sheet['A' + str(lineread)].value
                    sync.fullname = sheet['C' + str(lineread)].value
                    sync.current_user = sheet['D' + str(lineread)].value
                    sync.position = sheet['F' + str(lineread)].value
                    sync.department = sheet['E' + str(lineread)].value
                    sync.min = sheet['G' + str(lineread)].value
                    sync.plan = sheet['H' + str(lineread)].value
                    sync.package = sheet['I' + str(lineread)].value
                    sync.handset = sheet['J' + str(lineread)].value
                    sync.save()

        elif type_record == 'itequipment':
            sync = ItEquipmentModel()
            for lineread in range(2, (sheet.max_row + 1)):
                if sheet['A' + str(lineread)].value != None:
                    sync.pk = None
                    sync.folder = folderfk
                    sync.location1 = sheet['B' + str(lineread)].value
                    sync.location2 = sheet['C' + str(lineread)].value
                    sync.userfrom = sheet['D' + str(lineread)].value
                    sync.custodian = sheet['E' + str(lineread)].value
                    sync.device = sheet['F' + str(lineread)].value
                    sync.mobile = sheet['G' + str(lineread)].value
                    sync.imei = sheet['H' + str(lineread)].value
                    sync.iccid = sheet['I' + str(lineread)].value
                    sync.accessories = sheet['J' + str(lineread)].value
                    sync.usertype = sheet['K' + str(lineread)].value
                    sync.operation_manager = sheet['L' + str(lineread)].value
                    sync.save()

        messages.success(self.request, 'Synced successfully.')

        return super(UploadClassBaseView, self).form_valid(form)

    def form_invalid(self, form):
        messages.warning(self.request, 'Sorry I can not process your uploading as of now, Try again.')
        return self.response_to_render(self.get_context_data(form = form))

    def get_success_url(self):
        return reverse_lazy('core:upload')

class FolderClassBaseView(ListView):
    template_name = 'components/crud/view.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_view'] = 'folder'
        kwargs['context_title'] = 'Folders'
        return super(FolderClassBaseView, self).get_context_data(**kwargs)

    def get_queryset(self):
        return FolderModel.objects.filter(type_record = self.kwargs['type'])

class RecordClassBaseView(ListView):
    template_name = 'components/crud/view.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        kwargs['context_view'] = 'record'
        kwargs['context_folder'] = self.kwargs['folder']
        kwargs['context_type'] = self.kwargs['type']
        kwargs['context_title'] = 'Record'
        return super(RecordClassBaseView, self).get_context_data(**kwargs)

    def get_queryset(self):
        return AdaModel.objects.filter(folder = self.kwargs['folder']) if self.kwargs['type'] == 'ada' else ItEquipmentModel.objects.filter(folder = self.kwargs['folder'])

class FolderUpdateClassBaseView(UpdateView):
    template_name = 'components/crud/edit.html'
    form_class = FolderForm

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Modify Folder'
        return super(FolderUpdateClassBaseView, self).get_context_data(**kwargs)

    def get_queryset(self):
        return FolderModel.objects.filter(pk = self.kwargs[self.pk_url_kwarg])

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'Successfully modified.')
        return super(FolderUpdateClassBaseView, self).form_valid(form)

    def get_success_url(self):
        context = {}
        context['pk'] = self.kwargs[self.pk_url_kwarg]
        return reverse_lazy('core:updatefolder', kwargs = context)

class AdaUpdateClassBaseView(UpdateView):
    template_name = 'components/crud/edit.html'
    form_class = AdaForm

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Modify ADA Record'
        return super(AdaUpdateClassBaseView, self).get_context_data(**kwargs)

    def get_queryset(self):
        return AdaModel.objects.filter(pk = self.kwargs[self.pk_url_kwarg])

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'Successfully modify data record.')
        return super(AdaUpdateClassBaseView, self).form_valid(form)
    
    def get_success_url(self):
        context = {}
        context['pk'] = self.kwargs[self.pk_url_kwarg]
        return reverse_lazy('core:updateada', kwargs = context)

class ItEquipmentUpdateClassBaseView(UpdateView):
    template_name = 'components/crud/edit.html'
    form_class = ItEquipmentForm

    def get_context_data(self, **kwargs):
        kwargs['context_title'] = 'Modify IT Equipment Record'
        return super(ItEquipmentUpdateClassBaseView, self).get_context_data(**kwargs)

    def get_queryset(self):
        return ItEquipmentModel.objects.filter(pk = self.kwargs[self.pk_url_kwarg])

    def form_valid(self, form):
        self.object = form.save()
        messages.success(self.request, 'Successfully modify data record.')
        return super(ItEquipmentUpdateClassBaseView, self).form_valid(form)
    
    def get_success_url(self):
        context = {}
        context['pk'] = self.kwargs[self.pk_url_kwarg]
        return reverse_lazy('core:updateit', kwargs = context)

class FolderDeleteClassBaseView(DeleteView):
    model = FolderModel
    template_name = None

    def get_success_url(self):
        context = {}
        context['type'] = self.kwargs['type']
        messages.success(self.request, 'Folder deleted.')
        return reverse_lazy('core:folder', kwargs = context)